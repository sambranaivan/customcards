import argparse
import re
import sqlite3
import unicodedata
from collections import OrderedDict
from typing import Dict, Iterable, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime guard for missing dependency
    load_workbook = None


MONSTER_ID_BASE = 920000000
SPELL_ID_BASE = 921000000
TRAP_ID_BASE = 922000000
SETCODE_BASE = 0x3000

ATTRIBUTE_MAP = {
    "EARTH": 0x1,
    "WATER": 0x2,
    "FIRE": 0x4,
    "WIND": 0x8,
    "LIGHT": 0x10,
    "DARK": 0x20,
    "DIVINE": 0x40,
}

RACE_MAP = {
    "WARRIOR": 0x1,
    "SPELLCASTER": 0x2,
    "FAIRY": 0x4,
    "FIEND": 0x8,
    "ZOMBIE": 0x10,
    "MACHINE": 0x20,
    "AQUA": 0x40,
    "PYRO": 0x80,
    "ROCK": 0x100,
    "WINGED BEAST": 0x200,
    "PLANT": 0x400,
    "INSECT": 0x800,
    "THUNDER": 0x1000,
    "DRAGON": 0x2000,
    "BEAST": 0x4000,
    "BEAST-WARRIOR": 0x8000,
    "DINOSAUR": 0x10000,
    "FISH": 0x20000,
    "SEA SERPENT": 0x40000,
    "REPTILE": 0x80000,
    "PSYCHIC": 0x100000,
    "DIVINE-BEAST": 0x200000,
    "CREATOR GOD": 0x400000,
    "WYRM": 0x800000,
    "CYBERSE": 0x1000000,
    "ILLUSION": 0x2000000,
}

MONSTER_EFFECT = 0x1 | 0x20
MONSTER_FUSION_EFFECT = 0x1 | 0x20 | 0x40
MONSTER_XYZ_EFFECT = 0x1 | 0x20 | 0x800000
MONSTER_LINK_EFFECT = 0x1 | 0x20 | 0x4000000

SPELL_TYPES = {
    "NORMAL SPELL": 0x2,
    "QUICK-PLAY SPELL": 0x2 | 0x10000,
    "FIELD SPELL": 0x2 | 0x80000,
    "EQUIP SPELL": 0x2 | 0x40000,
    "CONTINUOUS SPELL": 0x2 | 0x20000,
}

TRAP_TYPES = {
    "NORMAL TRAP": 0x4,
    "COUNTER TRAP": 0x4 | 0x100000,
    "CONTINUOUS TRAP": 0x4 | 0x20000,
}


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def fold_for_key(text: str) -> str:
    text = normalize_text(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def parse_level(raw_level: object) -> int:
    text = normalize_text(raw_level)
    if text in {"", "-", "—"}:
        return 0
    return int(float(text))


def parse_monster_type(card_type: object) -> int:
    text = fold_for_key(normalize_text(card_type))
    if text == "monstruo":
        return MONSTER_EFFECT
    if text in {"fusion", "fusión"}:
        return MONSTER_FUSION_EFFECT
    if text.startswith("xyz rango"):
        return MONSTER_XYZ_EFFECT
    if text.startswith("link-"):
        return MONSTER_LINK_EFFECT
    raise ValueError(f"Tipo de monstruo no soportado: {card_type}")


def parse_spell_type(spell_type: object) -> int:
    key = normalize_text(spell_type).upper()
    if key not in SPELL_TYPES:
        raise ValueError(f"Tipo de spell no soportado: {spell_type}")
    return SPELL_TYPES[key]


def parse_trap_type(trap_type: object) -> int:
    key = normalize_text(trap_type).upper()
    if key not in TRAP_TYPES:
        raise ValueError(f"Tipo de trap no soportado: {trap_type}")
    return TRAP_TYPES[key]


def parse_link_rating(card_type: object) -> int:
    match = re.search(r"Link-(\d+)", normalize_text(card_type), flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def parse_xyz_rank(card_type: object) -> int:
    match = re.search(r"Xyz\s*Rango\s*(\d+)", normalize_text(card_type), flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def parse_archetype_tags(raw: object) -> List[str]:
    return [tag.strip() for tag in re.findall(r"\[([^\]]+)\]", normalize_text(raw))]


def pack_setcodes(values: Iterable[int]) -> int:
    packed = 0
    for idx, value in enumerate(values):
        if idx >= 4:
            raise ValueError("Solo se soportan hasta 4 setcodes por carta.")
        packed |= (int(value) & 0xFFFF) << (idx * 16)
    return packed


def monster_id_for_index(index: int) -> int:
    return MONSTER_ID_BASE + index


def spell_id_for_index(index: int) -> int:
    return SPELL_ID_BASE + index


def trap_id_for_index(index: int) -> int:
    return TRAP_ID_BASE + index


def build_setcode_map(rows: Iterable[Tuple[str, str]]) -> Dict[str, int]:
    tag_original_by_key = {}
    for _, raw_archetypes in rows:
        for tag in parse_archetype_tags(raw_archetypes):
            key = fold_for_key(tag)
            if key and key not in tag_original_by_key:
                tag_original_by_key[key] = tag

    setcode_map: Dict[str, int] = {}
    for offset, key in enumerate(sorted(tag_original_by_key)):
        setcode_map[key] = SETCODE_BASE + offset
    return setcode_map


def parse_attribute(raw: object) -> int:
    key = normalize_text(raw).upper()
    if not key:
        return 0
    if key not in ATTRIBUTE_MAP:
        raise ValueError(f"Atributo no soportado: {raw}")
    return ATTRIBUTE_MAP[key]


def parse_race(raw: object) -> int:
    key = normalize_text(raw).upper()
    if not key:
        return 0
    if key not in RACE_MAP:
        raise ValueError(f"Raza no soportada: {raw}")
    return RACE_MAP[key]


def to_texts_row(card_id: int, name: str, description: str) -> Tuple[object, ...]:
    return (card_id, name, description, *[""] * 16)


def to_setcode(raw_archetypes: object, setcode_map: Dict[str, int]) -> int:
    seen = OrderedDict()
    for tag in parse_archetype_tags(raw_archetypes):
        key = fold_for_key(tag)
        if key and key in setcode_map:
            seen[key] = setcode_map[key]
    return pack_setcodes(seen.values())


def read_sheet_rows(ws, min_col: int, max_col: int) -> List[Tuple[object, ...]]:
    rows = []
    for row in ws.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True):
        if normalize_text(row[1]):
            rows.append(row)
    return rows


def import_cards(excel_path: str, db_path: str) -> None:
    if load_workbook is None:
        raise RuntimeError("openpyxl no esta instalado. Instala con: pip install openpyxl")

    wb = load_workbook(excel_path, data_only=True)
    monsters = read_sheet_rows(wb["Monstruos"], 1, 12)
    spells = read_sheet_rows(wb["Spells"], 1, 6)
    traps = read_sheet_rows(wb["Traps"], 1, 6)

    if len(monsters) != 239 or len(spells) != 102 or len(traps) != 38:
        raise ValueError(
            f"Conteos inesperados: monstruos={len(monsters)}, spells={len(spells)}, traps={len(traps)}"
        )

    setcode_rows = []
    setcode_rows.extend((row[1], row[9]) for row in monsters)
    setcode_rows.extend((row[1], row[4]) for row in spells)
    setcode_rows.extend((row[1], row[4]) for row in traps)
    setcode_map = build_setcode_map(setcode_rows)

    datas_rows = []
    texts_rows = []

    for i, row in enumerate(monsters, start=1):
        card_id = monster_id_for_index(i)
        name = normalize_text(row[1])
        card_type_text = normalize_text(row[8])
        if card_type_text.lower().startswith("link-"):
            level = parse_link_rating(card_type_text)
        elif card_type_text.lower().startswith("xyz"):
            level = parse_xyz_rank(card_type_text)
        else:
            level = parse_level(row[3])
        datas_rows.append(
            (
                card_id,
                4,
                0,
                to_setcode(row[9], setcode_map),
                parse_monster_type(card_type_text),
                int(float(row[6] or 0)),
                int(float(row[7] or 0)),
                level,
                parse_race(row[5]),
                parse_attribute(row[4]),
                0,
            )
        )
        texts_rows.append(to_texts_row(card_id, name, normalize_text(row[11])))

    for i, row in enumerate(spells, start=1):
        card_id = spell_id_for_index(i)
        datas_rows.append(
            (
                card_id,
                4,
                0,
                to_setcode(row[4], setcode_map),
                parse_spell_type(row[3]),
                0,
                0,
                0,
                0,
                0,
                0,
            )
        )
        texts_rows.append(to_texts_row(card_id, normalize_text(row[1]), normalize_text(row[5])))

    for i, row in enumerate(traps, start=1):
        card_id = trap_id_for_index(i)
        datas_rows.append(
            (
                card_id,
                4,
                0,
                to_setcode(row[4], setcode_map),
                parse_trap_type(row[3]),
                0,
                0,
                0,
                0,
                0,
                0,
            )
        )
        texts_rows.append(to_texts_row(card_id, normalize_text(row[1]), normalize_text(row[5])))

    conn = sqlite3.connect(db_path)
    try:
        c = conn.cursor()
        c.executemany("INSERT OR REPLACE INTO datas VALUES (?,?,?,?,?,?,?,?,?,?,?)", datas_rows)
        c.executemany("INSERT OR REPLACE INTO texts VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", texts_rows)
        conn.commit()
    finally:
        conn.close()

    print(f"Monstruos importados: {len(monsters)}")
    print(f"Spells importadas: {len(spells)}")
    print(f"Traps importadas: {len(traps)}")
    print(f"Arquetipos detectados: {len(setcode_map)}")
    print(f"Total cartas importadas: {len(datas_rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa set Pokemon desde Excel a una base .cdb")
    parser.add_argument(
        "--excel",
        default="Pokemon_YGO_Todas_Las_Cartas.xlsx",
        help="Ruta al archivo Excel fuente",
    )
    parser.add_argument(
        "--db",
        default="expansions/cards-unofficial.cdb",
        help="Ruta a la base .cdb destino",
    )
    args = parser.parse_args()
    import_cards(args.excel, args.db)


if __name__ == "__main__":
    main()
